from tkinter import *
from tkinter.filedialog import askopenfilename
import pandas as pd
import numpy as np
import datetime
import openpyxl
import re
from fuzzywuzzy import fuzz
import os
import mysql.connector
import math

base = Tk()
base.title('Создание ведомости участников')
base.geometry('820x720')
base.resizable(width=True, height=True)
check_fio = BooleanVar()
uniq_create = BooleanVar()
create_41 = BooleanVar()
create_41.set(True)
f_left = Frame(base)
f_right = Frame(base)
logger = Text(f_right, width=85, height=40, bg="white", fg='blue', wrap=WORD)
scroll = Scrollbar(command=logger.yview)
scroll.pack(side=LEFT, fill=Y)
logger.config(yscrollcommand=scroll.set)
checkbutton = Checkbutton(f_left, text="Исправлять ФИО", variable=check_fio, onvalue=True, offvalue=False, padx=5,
                          pady=1)
checkbutton_create_uniq = Checkbutton(f_left, text="Создавать файл уникальных участников", variable=uniq_create,
                                      onvalue=True, offvalue=False, padx=5, pady=1)
checkbutton_41_60 = Checkbutton(f_left, text="Учет всех по определенному времени", variable=create_41,
                                      onvalue=True, offvalue=False, padx=5, pady=1)
values = {'Document': '-', 'Примечание': '-', 'Фирма изготовитель': '-'}
must_index = ['ФИО гражданина', 'Контактный телефон гражданина']
must_index0 = ['Идентификатор конференции', 'Тема']
name_list_all = []
time_list_all = []
time_format_in = "%H:%M"
time_format_in_secs = "%H:%M:%S"
list_of_not_used = ["куратор", "координатор", "оскад", "преподаватель", "организатор", "педагог", "iphone", "ipad"]
on_time = 41
off_time = 60
# Создаем словарь с месяцами
month_dict = {'01': 'Январь', '02': 'Февраль', '03': 'Март', '04': 'Апрель', '05': 'Май', '06': 'Июнь', '07': 'Июль',
              '08': 'Август', '09': 'Сентябрь', '10': 'Октябрь', '11': 'Ноябрь', '12': 'Декабрь'}

df0_ = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df1 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df2 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df3 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df4 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df5 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df6 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df7 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df8 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df9 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})
df10 = pd.DataFrame(
    {'Наименование': [], 'PART_NUMBER': [], 'Ref Des': [], 'Qty': [], 'Вариант': [], 'Примечание': [], 'Класс': [],
     'Фирма изготовитель': [], 'Функциональное назначение': []})


def file_opener_0(reset=False):
    global df0
    global group_name
    reset = reset
    if (reset == False):
        inpu0 = askopenfilename(initialdir="/", filetypes=[('Excel Files', '*.xlsx')])
    else:
        inpu0 = False
        l0['text'] = f'\n'
        df0 = pd.DataFrame()
        logger.insert(END, f'Сброс всех данных!\n')
    if (inpu0):
        logger.insert(END, f'Выбран файл основной ведомости {inpu0}\n')
        df0_ = pd.read_excel(inpu0)
        group_name = df0_.columns[0]
        l_group_name['text'] = group_name
        df0 = pd.read_excel(inpu0, skiprows=[0])
        df0 = df0.replace(np.nan, '0')
        # for x in must_index:
        #     if x not in df0.columns:
        #         df0 = []
        #         logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        df0['ФИО гражданина'] = df0[df0.columns[1]]
        df0['Контактный телефон гражданина'] = df0[df0.columns[2]]
        try:
            df0['Дата зачисления'] = df0[df0.columns[3]]
        except:
            df0['Дата зачисления'] = [0 for x in range(len(df0['ФИО гражданина'].tolist()))]
        df0['Контактный телефон гражданина'] = df0['Контактный телефон гражданина'].astype(str)
        df0['ФИО гражданина'] = df0['ФИО гражданина'].astype(str)
        try:
            df0['Дата зачисления'] = enrollment_date(df0['Дата зачисления'].tolist())
        except:
            pass

        df0 = df0[df0['ФИО гражданина'] != '0']
        df0['ФИО гражданина'] = df0['ФИО гражданина'].astype(str)
        df0['ФИО гражданина'] = first_obr(df0['ФИО гражданина'].tolist())
		# df0 = df0[df0['Контактный телефон гражданина'] != 0]
		# df0['Контактный телефон гражданина'] = first_obr(df0['Контактный телефон гражданина'].tolist())
		# df0['Контактный телефон гражданина'] = df0['Контактный телефон гражданина'].tolist()
        # Проверка на уникальность колонки ФИО
        if (df0['ФИО гражданина'].is_unique):
            logger.insert(END, f'Все ФИО ведомости уникальны!\n')
        else:
            logger.insert(END, f'\nОШИБКА!!!! ФИО ведомости не уникальны!\n')
        df0 = df0.reset_index()
		# df0.to_excel('df0.xlsx')
        l0['text'] = f'OK\n'


def file_opener_1(reset=False):
    global df1
    global df1_name
    global month_day
    global year
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df1 = pd.DataFrame()
        l1['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df1 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df1.columns:
                df1 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df1_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df1['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df1['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df1_name = f"{time_f_0} {time_f_1}"
        except:
            df1_name = f"{df1['Время начала'].iloc[0]} {df1['Время завершения'].iloc[0]}"
        # #################################################################
        # df1_name = f"{df1['Время начала'].iloc[0]} {df1['Время завершения'].iloc[0]}" # Время открытия и закрытия сессии zoom
        print(df1_name)
        df1_name = df1_name.split(' ')
        start = f'{df1_name[2]} {df1_name[1]}'  # Дата и время начала занятия
        end = f'{df1_name[2]} {df1_name[3]}'  # Дата и время завершения занятия
        date_for_column = df1_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        month_day = month_dict[date_for_column[0]]  # Вытаскиваем дату работает только для первой таблицы
        year = f'{date_for_column[2]}'  # год
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df1_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df1_time = df1['Продолжительность (минуты)'].iloc[0]
        df1.columns = df1.iloc[1]
        df1 = df1.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df1_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df1['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df1['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df1['Время входа'] = new_enter_time
            df1['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df1['Имя (настоящее имя)'] = first_obr(df1['Имя (настоящее имя)'].tolist())
        df1['Name'] = [df1_name for x in range(len(df1.index))]
        df1['Time'] = [df1_time for x in range(len(df1.index))]
        # df1.to_excel('df1.xlsx')
        df1['Secs'] = convert_to_sec(df1['Время входа'], df1['Время выхода'], start, end)
        name_list_all.append(df1_name)
        # time_list_all.append(df1_time)
        time_list_all.append(((datetime.datetime.strptime(en_end.get(), time_format_in) - datetime.datetime.strptime(
            en_start.get(), time_format_in)).total_seconds()) / 60)
        # df1.to_excel('df1.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df1['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        intersect_time_in_0 = []
        intersect_time_in_1 = []
        intersect_time_out_0 = []
        intersect_time_out_1 = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df1[df1['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df1[df1['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df1[df1['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num < num1):
                        in_time_1 = df1[df1['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df1[df1['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_time_in_0.append(in_time_0)
                            intersect_time_in_1.append(in_time_1)
                            intersect_time_out_0.append(out_time_0)
                            intersect_time_out_1.append(out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list,
             'Время входа 0': intersect_time_in_0, 'Время входа 1': intersect_time_in_1,
             'Время выхода 0': intersect_time_out_0, 'Время выхода 1': intersect_time_out_1})
        intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')
        # Чистим лишнее
        intersec_tab = find_intersections_same(intersec_tab, start, end)
        intersec_tab = find_intersections_same(intersec_tab, start, end)
        intersec_tab = find_intersections_same(intersec_tab, start, end)
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df1['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df1[df1['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
                        print(f'Делитель для {name}: {delitel}')
            else:
                common_delitel.append(0)
        print(common_delitel)
        df1['time_delta'] = common_delitel
        df1['Secs'] = df1['Secs'] - df1['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l1['text'] = f'OK\n'
    #df1.to_excel('df0.xlsx')
    # print(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(start_time, time_format_in)).total_seconds())/60)


def file_opener_2(reset=False):
    global df2
    global df2_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df2 = pd.DataFrame()
        l2['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df2 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df2.columns:
                df2 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df2_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df2['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df2['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df2_name = f"{time_f_0} {time_f_1}"
        except:
            df2_name = f"{df2['Время начала'].iloc[0]} {df2['Время завершения'].iloc[0]}"
        # #################################################################
        print(df2_name)
        df2_name = df2_name.split(' ')
        start = f'{df2_name[2]} {df2_name[1]}'  # Дата и время начала занятия
        end = f'{df2_name[2]} {df2_name[3]}'  # Дата и время завершения занятия
        date_for_column = df2_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df2_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df2_time = df2['Продолжительность (минуты)'].iloc[0]
        df2.columns = df2.iloc[1]
        df2 = df2.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df2_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df2['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df2['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df2['Время входа'] = new_enter_time
            df2['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df2['Имя (настоящее имя)'] = first_obr(df2['Имя (настоящее имя)'].tolist())
        df2['Name'] = [df2_name for x in range(len(df2.index))]
        df2['Time'] = [df2_time for x in range(len(df2.index))]
        df2['Secs'] = convert_to_sec(df2['Время входа'], df2['Время выхода'], start, end)
        name_list_all.append(df2_name)
        # time_list_all.append(df2_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df2.to_excel('df2.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df2['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df2[df2['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df2[df2['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df2[df2['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df2[df2['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df2[df2['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df2['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df2[df2['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df2['time_delta'] = common_delitel
        df2['Secs'] = df2['Secs'] - df2['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l2['text'] = f'OK\n'


def file_opener_3(reset=False):
    global df3
    global df3_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df3 = pd.DataFrame()
        l3['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df3 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df3.columns:
                df3 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df3_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df3['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df3['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df3_name = f"{time_f_0} {time_f_1}"
        except:
            df3_name = f"{df3['Время начала'].iloc[0]} {df3['Время завершения'].iloc[0]}"
        # #################################################################
        print(df3_name)
        df3_name = df3_name.split(' ')
        start = f'{df3_name[2]} {df3_name[1]}'  # Дата и время начала занятия
        end = f'{df3_name[2]} {df3_name[3]}'  # Дата и время завершения занятия
        date_for_column = df3_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df3_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df3_time = df3['Продолжительность (минуты)'].iloc[0]
        df3.columns = df3.iloc[1]
        df3 = df3.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df3_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df3['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df3['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df3['Время входа'] = new_enter_time
            df3['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df3['Имя (настоящее имя)'] = first_obr(df3['Имя (настоящее имя)'].tolist())
        df3['Name'] = [df3_name for x in range(len(df3.index))]
        df3['Time'] = [df3_time for x in range(len(df3.index))]
        df3['Secs'] = convert_to_sec(df3['Время входа'], df3['Время выхода'], start, end)
        name_list_all.append(df3_name)
        # time_list_all.append(df3_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df3.to_excel('df3.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df3['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df3[df3['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df3[df3['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df3[df3['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df3[df3['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df3[df3['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df3['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df3[df3['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df3['time_delta'] = common_delitel
        df3['Secs'] = df3['Secs'] - df3['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l3['text'] = f'OK\n'


def file_opener_4(reset=False):
    global df4
    global df4_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df4 = pd.DataFrame()
        l4['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df4 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df4.columns:
                df4 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df4_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df4['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df4['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df4_name = f"{time_f_0} {time_f_1}"
        except:
            df4_name = f"{df4['Время начала'].iloc[0]} {df4['Время завершения'].iloc[0]}"
        # #################################################################
        print(df4_name)
        df4_name = df4_name.split(' ')
        start = f'{df4_name[2]} {df4_name[1]}'  # Дата и время начала занятия
        end = f'{df4_name[2]} {df4_name[3]}'  # Дата и время завершения занятия
        date_for_column = df4_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df4_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df4_time = df4['Продолжительность (минуты)'].iloc[0]
        df4.columns = df4.iloc[1]
        df4 = df4.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df4_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df4['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df4['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df4['Время входа'] = new_enter_time
            df4['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df4['Имя (настоящее имя)'] = first_obr(df4['Имя (настоящее имя)'].tolist())
        df4['Name'] = [df4_name for x in range(len(df4.index))]
        df4['Time'] = [df4_time for x in range(len(df4.index))]
        df4['Secs'] = convert_to_sec(df4['Время входа'], df4['Время выхода'], start, end)
        name_list_all.append(df4_name)
        # time_list_all.append(df4_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df4.to_excel('df4.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df4['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df4[df4['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df4[df4['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df4[df4['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df4[df4['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df4[df4['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df4['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df4[df4['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df4['time_delta'] = common_delitel
        df4['Secs'] = df4['Secs'] - df4['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l4['text'] = f'OK\n'


def file_opener_5(reset=False):
    global df5
    global df5_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df5 = pd.DataFrame()
        l5['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df5 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df5.columns:
                df5 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df5_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df5['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df5['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df5_name = f"{time_f_0} {time_f_1}"
        except:
            df5_name = f"{df5['Время начала'].iloc[0]} {df5['Время завершения'].iloc[0]}"
        # #################################################################
        print(df5_name)
        df5_name = df5_name.split(' ')
        start = f'{df5_name[2]} {df5_name[1]}'  # Дата и время начала занятия
        end = f'{df5_name[2]} {df5_name[3]}'  # Дата и время завершения занятия
        date_for_column = df5_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df5_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df5_time = df5['Продолжительность (минуты)'].iloc[0]
        df5.columns = df5.iloc[1]
        df5 = df5.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df5_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df5['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df5['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df5['Время входа'] = new_enter_time
            df5['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df5['Имя (настоящее имя)'] = first_obr(df5['Имя (настоящее имя)'].tolist())
        df5['Name'] = [df5_name for x in range(len(df5.index))]
        df5['Time'] = [df5_time for x in range(len(df5.index))]
        df5['Secs'] = convert_to_sec(df5['Время входа'], df5['Время выхода'], start, end)
        name_list_all.append(df5_name)
        # time_list_all.append(df5_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df5.to_excel('df5.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df5['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df5[df5['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df5[df5['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df5[df5['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df5[df5['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df5[df5['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df5['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df5[df5['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df5['time_delta'] = common_delitel
        df5['Secs'] = df5['Secs'] - df5['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l5['text'] = f'OK\n'


def file_opener_6(reset=False):
    global df6
    global df6_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df6 = pd.DataFrame()
        l6['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df6 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df6.columns:
                df6 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df6_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df6['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df6['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df6_name = f"{time_f_0} {time_f_1}"
        except:
            df6_name = f"{df6['Время начала'].iloc[0]} {df6['Время завершения'].iloc[0]}"
        # #################################################################
        print(df6_name)
        df6_name = df6_name.split(' ')
        start = f'{df6_name[2]} {df6_name[1]}'  # Дата и время начала занятия
        end = f'{df6_name[2]} {df6_name[3]}'  # Дата и время завершения занятия
        date_for_column = df6_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df6_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df6_time = df6['Продолжительность (минуты)'].iloc[0]
        df6.columns = df6.iloc[1]
        df6 = df6.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df6_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df6['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df6['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df6['Время входа'] = new_enter_time
            df6['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df6['Имя (настоящее имя)'] = first_obr(df6['Имя (настоящее имя)'].tolist())
        df6['Name'] = [df6_name for x in range(len(df6.index))]
        df6['Time'] = [df6_time for x in range(len(df6.index))]
        df6['Secs'] = convert_to_sec(df6['Время входа'], df6['Время выхода'], start, end)
        name_list_all.append(df6_name)
        # time_list_all.append(df6_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df6.to_excel('df6.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df6['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df6[df6['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df6[df6['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df6[df6['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df6[df6['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df6[df6['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df6['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df6[df6['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df6['time_delta'] = common_delitel
        df6['Secs'] = df6['Secs'] - df6['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l6['text'] = f'OK\n'


def file_opener_7(reset=False):
    global df7
    global df7_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df7 = pd.DataFrame()
        l7['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df7 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df7.columns:
                df7 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df7_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df7['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df7['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df7_name = f"{time_f_0} {time_f_1}"
        except:
            df7_name = f"{df7['Время начала'].iloc[0]} {df7['Время завершения'].iloc[0]}"
        # #################################################################
        print(df7_name)
        df7_name = df7_name.split(' ')
        start = f'{df7_name[2]} {df7_name[1]}'  # Дата и время начала занятия
        end = f'{df7_name[2]} {df7_name[3]}'  # Дата и время завершения занятия
        date_for_column = df7_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df7_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df7_time = df7['Продолжительность (минуты)'].iloc[0]
        df7.columns = df7.iloc[1]
        df7 = df7.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df7_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df7['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df7['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df7['Время входа'] = new_enter_time
            df7['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df7['Имя (настоящее имя)'] = first_obr(df7['Имя (настоящее имя)'].tolist())
        df7['Name'] = [df7_name for x in range(len(df7.index))]
        df7['Time'] = [df7_time for x in range(len(df7.index))]
        df7['Secs'] = convert_to_sec(df7['Время входа'], df7['Время выхода'], start, end)
        name_list_all.append(df7_name)
        # time_list_all.append(df7_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df7.to_excel('df7.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df7['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df7[df7['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df7[df7['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df7[df7['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df7[df7['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df7[df7['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df7['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df7[df7['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df7['time_delta'] = common_delitel
        df7['Secs'] = df7['Secs'] - df7['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l7['text'] = f'OK\n'


def file_opener_8(reset=False):
    global df8
    global df8_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df8 = pd.DataFrame()
        l8['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df8 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df8.columns:
                df8 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df8_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df8['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df8['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df8_name = f"{time_f_0} {time_f_1}"
        except:
            df8_name = f"{df8['Время начала'].iloc[0]} {df8['Время завершения'].iloc[0]}"
        # #################################################################
        print(df8_name)
        df8_name = df8_name.split(' ')
        start = f'{df8_name[2]} {df8_name[1]}'  # Дата и время начала занятия
        end = f'{df8_name[2]} {df8_name[3]}'  # Дата и время завершения занятия
        date_for_column = df8_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df8_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df8_time = df8['Продолжительность (минуты)'].iloc[0]
        df8.columns = df8.iloc[1]
        df8 = df8.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df8_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df8['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df8['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df8['Время входа'] = new_enter_time
            df8['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df8['Имя (настоящее имя)'] = first_obr(df8['Имя (настоящее имя)'].tolist())
        df8['Name'] = [df8_name for x in range(len(df8.index))]
        df8['Time'] = [df8_time for x in range(len(df8.index))]
        df8['Secs'] = convert_to_sec(df8['Время входа'], df8['Время выхода'], start, end)
        name_list_all.append(df8_name)
        # time_list_all.append(df8_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df8.to_excel('df8.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df8['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df8[df8['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df8[df8['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df8[df8['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df8[df8['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df8[df8['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df8['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df8[df8['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df8['time_delta'] = common_delitel
        df8['Secs'] = df8['Secs'] - df8['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l8['text'] = f'OK\n'


def file_opener_9(reset=False):
    global df9
    global df9_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df9 = pd.DataFrame()
        l9['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df9 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df9.columns:
                df9 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df9_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df9['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df9['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df9_name = f"{time_f_0} {time_f_1}"
        except:
            df9_name = f"{df9['Время начала'].iloc[0]} {df9['Время завершения'].iloc[0]}"
        # #################################################################
        print(df9_name)
        df9_name = df9_name.split(' ')
        start = f'{df9_name[2]} {df9_name[1]}'  # Дата и время начала занятия
        end = f'{df9_name[2]} {df9_name[3]}'  # Дата и время завершения занятия
        date_for_column = df9_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df9_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df9_time = df9['Продолжительность (минуты)'].iloc[0]
        df9.columns = df9.iloc[1]
        df9 = df9.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df9_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df9['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df9['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df9['Время входа'] = new_enter_time
            df9['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df9['Имя (настоящее имя)'] = first_obr(df9['Имя (настоящее имя)'].tolist())
        df9['Name'] = [df9_name for x in range(len(df9.index))]
        df9['Time'] = [df9_time for x in range(len(df9.index))]
        df9['Secs'] = convert_to_sec(df9['Время входа'], df9['Время выхода'], start, end)
        name_list_all.append(df9_name)
        # time_list_all.append(df9_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df9.to_excel('df9.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df9['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df9[df9['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df9[df9['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df9[df9['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df9[df9['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df9[df9['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')

        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df9['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df9[df9['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df9['time_delta'] = common_delitel
        df9['Secs'] = df9['Secs'] - df9['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l9['text'] = f'OK\n'


def file_opener_10(reset=False):
    global df10
    global df10_name
    reset = reset
    if (reset == False):
        inpu1 = askopenfilename(initialdir="/", filetypes=[('CSV Files', '*.csv')])
    else:
        inpu1 = False
        df10 = pd.DataFrame()
        l10['text'] = f'\n'
    if (inpu1):
        # Считываем данные из поля времени:
        start_time = en_start.get()
        start_time = f"{start_time}"  # время начала в формате 00:00
        end_time = en_end.get()
        end_time = f"{end_time}"  # время конца в формате 00:00
        ##############################################################
        logger.insert(END, f'Выбран файл {inpu1}\n')
        df10 = pd.read_csv(inpu1)
        for x in must_index0:
            if x not in df10.columns:
                df10 = []
                logger.insert(END, f'ERROR: Выбраный файл не соответствует формату!!!\n')
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df10_name
        try:
            time_f_0 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df10['Время начала'].iloc[0])
            if (len(time_f_0) > 0):
                time_f_0 = time_f_0[0].split(' ')
                time_f_0 = f"{time_f_0[0].split('.')[1]}/{time_f_0[0].split('.')[0]}/{time_f_0[0].split('.')[2]} {time_f_0[1]}"
            else:
                raise Exception("Date format except")
            time_f_1 = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', df10['Время завершения'].iloc[0])
            if (len(time_f_1) > 0):
                time_f_1 = time_f_1[0].split(' ')
                time_f_1 = f"{time_f_1[0].split('.')[1]}/{time_f_1[0].split('.')[0]}/{time_f_1[0].split('.')[2]} {time_f_1[1]}"
            df10_name = f"{time_f_0} {time_f_1}"
        except:
            df10_name = f"{df10['Время начала'].iloc[0]} {df10['Время завершения'].iloc[0]}"
        # #################################################################
        print(df10_name)
        df10_name = df10_name.split(' ')
        start = f'{df10_name[2]} {df10_name[1]}'  # Дата и время начала занятия
        end = f'{df10_name[2]} {df10_name[3]}'  # Дата и время завершения занятия
        date_for_column = df10_name[0].split('/')  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        date_for_column = f'{date_for_column[1]}.{date_for_column[0]}.{date_for_column[2][2:4]}'  # Дата проведения занятия в формате для ведомости дд.мм.гг.
        df10_name = f'''Дата
({date_for_column})
Время проведения занятия
({start_time}-{end_time})'''
        df10_time = df10['Продолжительность (минуты)'].iloc[0]
        df10.columns = df10.iloc[1]
        df10 = df10.drop([0, 1])
        # Для второго заказчика необходим другой формат времение
        # дд.мм.гг - у Юли мм/дд/гг, соответствено переведем все в один формат, если это необходимо.df10_name
        try:
            new_enter_time = []
            new_exit_time = []
            for time in df10['Время входа'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_enter_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            for time in df10['Время выхода'].tolist():
                time_f = re.findall(r'\d\d[.]\d\d[.]\d\d\d\d \d\d[:]\d\d[:]\d\d', time)
                time_f = time_f[0].split(" ")
                new_exit_time.append(
                    f'{time_f[0].split(".")[1]}/{time_f[0].split(".")[0]}/{time_f[0].split(".")[2]} {time_f[1]}')
            df10['Время входа'] = new_enter_time
            df10['Время выхода'] = new_exit_time
        except:
            pass
        ### Конец перевода формата ###
        df10['Имя (настоящее имя)'] = first_obr(df10['Имя (настоящее имя)'].tolist())
        df10['Name'] = [df10_name for x in range(len(df10.index))]
        df10['Time'] = [df10_time for x in range(len(df10.index))]
        df10['Secs'] = convert_to_sec(df10['Время входа'], df10['Время выхода'], start, end)
        name_list_all.append(df10_name)
        # time_list_all.append(df10_time)
        time_list_all.append(((datetime.datetime.strptime(end_time, time_format_in) - datetime.datetime.strptime(
            start_time, time_format_in)).total_seconds()) / 60)
        # df10.to_excel('df10.xlsx')
        # Необходимо найти пересечения по времени для конкретного пользователя (Когда с разных аккаунтов
        # заходят пользователи)
        uniq_name = pd.Series(df10['Имя (настоящее имя)'].tolist()).unique()
        intersect_name = []
        intersect_time = []
        time_delta_list = []
        for name in uniq_name:
            iters_for_name = len(df10[df10['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
            for num in range(iters_for_name):
                in_time_0 = df10[df10['Имя (настоящее имя)'] == name].iloc[num]['Время входа']
                out_time_0 = df10[df10['Имя (настоящее имя)'] == name].iloc[num]['Время выхода']
                for num1 in range(iters_for_name):
                    if (num != num1) & (num < num1):
                        in_time_1 = df10[df10['Имя (настоящее имя)'] == name].iloc[num1]['Время входа']
                        out_time_1 = df10[df10['Имя (настоящее имя)'] == name].iloc[num1]['Время выхода']
                        date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                        if (date_inersec[0]):
                            # print(in_time_0, out_time_0, in_time_1, out_time_1)
                            intersect_name.append(name)
                            intersect_time.append(
                                f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                            time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list})
        # intersec_tab = intersec_tab.drop_duplicates()
        logger.insert(END, f'\nДанные ФИО из таблицы {inpu1} имеют параллельные входы в аккаунт:\n')
        for intersec_fio, intersec_time, delta in zip(intersec_tab['ФИО'].tolist(),
                                                      intersec_tab['Время пересечения'].tolist(),
                                                      intersec_tab['Time Delta'].tolist()):
            logger.insert(END, f'\n{intersec_fio} ({delta} секунд)\n{intersec_time}\n')
        # Чистим лишнее
        intersec_tab = intersec_tab.groupby(['ФИО'])['Time Delta'].sum()
        print(intersec_tab)
        intersec_tab = intersec_tab.reset_index()
        common_delitel = []
        for number, name in enumerate(df10['Имя (настоящее имя)'].tolist()):
            if name in intersec_tab['ФИО'].tolist():
                for number1, name1 in enumerate(intersec_tab['ФИО'].tolist()):
                    if (name == name1):
                        iters_for_name = len(df10[df10['Имя (настоящее имя)'] == name]['Имя (настоящее имя)'])
                        delitel = int(intersec_tab.iloc[number1]['Time Delta']) / iters_for_name
                        common_delitel.append(delitel)
            else:
                common_delitel.append(0)
        print(common_delitel)
        df10['time_delta'] = common_delitel
        df10['Secs'] = df10['Secs'] - df10['time_delta']
        # Конец блока предварительного поиска пересечения по времени
        l10['text'] = f'OK\n'


def time_equal(enrollment_date, date_zaniat):
    time_format = "%d.%m.%Y"  # %H:%M:%S     %m/%d/%Y %H:%M:%S
    try:
        print(enrollment_date)
        print(date_zaniat)
        enrollment_date = datetime.datetime.strptime(enrollment_date, time_format)
        date_zaniat = datetime.datetime.strptime(date_zaniat, time_format)
        razn = (enrollment_date - date_zaniat).total_seconds()
        if razn <= 0:
            return True
        else:
            return False
    except:
        return True


def convert_summ_date(date):
    date_list = []
    for x in date:
        list_date = re.findall(r'\d\d[.]\d\d[.]\d\d', x)
        list_date = list_date[0].split('.')
        month = list_date[1]
        day = list_date[0]
        year = list_date[2]
        date_list.append(f'{day}.{month}.20{year}')
    return date_list


def enrollment_date(date):
    date_list = []
    for x in date:
        try:
            list_date = x.split("/")
            month = list_date[0]
            if (len(month) < 2):
                month = f'0{month}'
            day = list_date[1]
            if (len(day) < 2):
                day = f'0{day}'
            year = list_date[2]
            date_list.append(f'{day}.{month}.{year}')
        except:
            try:
                print(str(x.date()).split('-'))
                list_date1 = str(x.date()).split('-')
                month = list_date1[1]
                if (len(month) < 2):
                    month = f'0{month}'
                day = list_date1[2]
                if (len(day) < 2):
                    day = f'0{day}'
                year = list_date1[0]
                date_list.append(f'{day}.{month}.{year}')
            except:
                date_list.append(np.nan)
    print(date_list)
    return date_list


def date_intersection(t1start, t1end, t2start, t2end):
    if t1end < t2start: return False, False
    if t1end == t2start: return False, False  #
    if t1start == t2start: return True, t1start, t1end
    if t1start < t2start and t2start < t1end and t1end > t2end: return True, t1start, t1end
    if t1start < t2start and t2start < t1end and t1end < t2end: return True, t1start, t2end
    if t1start > t2start and t1end < t2end: return True, t2start, t2end
    if t1start < t2start and t1end > t2end: return True, t1start, t1end
    if t1start < t2end and t1end > t2end: return True, t1start, t1end
    if t1start > t2start and t1start < t2end: return True, t2start, t2end
    if t1start == t2end: return False, False  #
    if t1end == t2end: return True, t1start, t1end
    if t1start > t2end: return False, False


def first_obr(fio):
    new_list = []
    fio_list = fio
    for x in fio_list:
        temp = x
        print(x)
        x = str(x)
        x = x.replace('  ', ' ')
        x = x.replace('😂', '')
        x = x.replace('\xa0', ' ')
        x = x.replace('Ё', 'Е')
        x = x.replace('ё', 'е')
        x = x.replace('0', '')
        x = x.replace('1', '')
        x = x.replace('2', '')
        x = x.replace('3', '')
        x = x.replace('4', '')
        x = x.replace('5', '')
        x = x.replace('6', '')
        x = x.replace('7', '')
        x = x.replace('8', '')
        x = x.replace('9', '')
        x = x.replace('#', '')
        x = x.replace('!', '')
        x = x.replace('?', '')
        x = x.replace('.', '')
        x = x.replace(',', '')
        x = x.lstrip()
        x = x.rstrip()
        x = first_obr_zoom(x)
        x = x.title()
        if len(x) < 2:
            logger.insert(END, f"{temp} ФИО возможно содержит ошибки или пустое\n")
            x = 'Пустое ФИО'
        if len(x) <= 2:
            logger.insert(END, f"{temp} ФИО возможно содержит ошибки\n")
            x = 7 * x
        elif len(x) < 5:
            logger.insert(END, f"{temp} ФИО возможно содержит ошибки\n")
            x = 4 * x
        print(x)
        new_list.append(x)
    return new_list


def first_obr_zoom(fio):  # Входная переменная строка ФИО
    fio = fio
    new_list = []
    fio = fio.replace('  ', ' ')
    # fio = fio.title()
    fio = fio.lstrip()
    fio = fio.rstrip()
    fio_list = fio.split(' ')
    if (len(fio_list) > 3):
        fio = f'{fio_list[0]} {fio_list[1]} {fio_list[2]}'
    a = re.findall('[А-Я][а-я]\w+[А-Я]', fio)
    if a:
        a = list(a)[0]
        a = a[0:(len(a) - 1)]
        fio = a + ' ' + fio.replace(a, '')
    fio_list = fio.split(' ')
    if (len(fio_list) > 3):
        fio = f'{fio_list[0]} {fio_list[1]} {fio_list[2]}'

    return fio


def reset_functions():
    file_opener_0(reset=True)
    file_opener_1(reset=True)
    file_opener_2(reset=True)
    file_opener_3(reset=True)
    file_opener_4(reset=True)
    file_opener_5(reset=True)
    file_opener_6(reset=True)
    file_opener_7(reset=True)
    file_opener_8(reset=True)
    file_opener_9(reset=True)
    file_opener_10(reset=True)


def time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end):
    in_time_0 = in_time_0
    out_time_0 = out_time_0
    in_time_1 = in_time_1
    out_time_1 = out_time_1

    time_1_hour = en_check_time_1.get()
    time_1_hour_start = int(time_1_hour.split('-')[0].lstrip().rstrip())
    time_1_hour_end = int(time_1_hour.split('-')[1].lstrip().rstrip())

    time_2_hour = en_check_time_2.get()
    time_2_hour_start = int(time_2_hour.split('-')[0].lstrip().rstrip())
    time_2_hour_end = int(time_2_hour.split('-')[1].lstrip().rstrip())

    lesson_time = ((datetime.datetime.strptime(en_end.get(), time_format_in) - datetime.datetime.strptime(
        en_start.get(), time_format_in)).total_seconds()) / 60

    if (create_41.get()):
        if (lesson_time >= 110):
            start_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_2_hour_start)
            start_time = str(start_time)
            start_time = start_time.split(' ')[1]
            start_time = f"{start.split(' ')[0]} {start_time}" # время начала в формате 00:00
            end_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_2_hour_end)
            end_time = str(end_time)
            end_time = end_time.split(' ')[1]
            end_time = f"{end.split(' ')[0]} {end_time}"
        elif (lesson_time >= 50):
            start_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_1_hour_start)
            start_time = str(start_time)
            start_time = start_time.split(' ')[1]
            start_time = f"{start.split(' ')[0]} {start_time}" # время начала в формате 00:00
            end_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_1_hour_end)
            end_time = str(end_time)
            end_time = end_time.split(' ')[1]
            end_time = f"{end.split(' ')[0]} {end_time}"
    else:
        start_time = en_start.get()
        start_time = f"{start.split(' ')[0]} {start_time}:00"
        end_time = en_end.get()
        end_time = f"{end.split(' ')[0]} {end_time}:00"

    time_format = "%m/%d/%Y %H:%M:%S"  # %H:%M:%S     %m/%d/%Y %H:%M:%S

    start_str = datetime.datetime.strptime(start_time, time_format)
    end_str = datetime.datetime.strptime(end_time, time_format)
    in_time_0_str = datetime.datetime.strptime(in_time_0, time_format)
    out_time_0_str = datetime.datetime.strptime(out_time_0, time_format)
    in_time_1_str = datetime.datetime.strptime(in_time_1, time_format)
    out_time_1_str = datetime.datetime.strptime(out_time_1, time_format)

    if (in_time_0_str < start_str):
        in_time_0 = start_time
    if (in_time_0_str > end_str):
        in_time_0 = end_time

    if (out_time_0_str < start_str):
        out_time_0 = start_time
    if (out_time_0_str > end_str):
        out_time_0 = end_time

    if (in_time_1_str < start_str):
        in_time_1 = start_time
    if (in_time_1_str > end_str):
        in_time_1 = end_time

    if (out_time_1_str < start_str):
        out_time_1 = start_time
    if (out_time_1_str > end_str):
        out_time_1 = end_time

    first_delta = (datetime.datetime.strptime(out_time_0, time_format) - datetime.datetime.strptime(in_time_0,
                                                                                                    time_format)).total_seconds()
    second_delta = (datetime.datetime.strptime(out_time_1, time_format) - datetime.datetime.strptime(in_time_1,
                                                                                                     time_format)).total_seconds()
    # delta = abs(first_delta - second_delta) # Время

    delta_in = (datetime.datetime.strptime(in_time_0, time_format) - datetime.datetime.strptime(in_time_1,
                                                                                                time_format)).total_seconds()
    delta_out = (datetime.datetime.strptime(out_time_0, time_format) - datetime.datetime.strptime(out_time_1,
                                                                                                  time_format)).total_seconds()

    if (delta_in <= 0) & (delta_out <= 0):
        delta = (datetime.datetime.strptime(out_time_0, time_format) - datetime.datetime.strptime(in_time_1,
                                                                                                  time_format)).total_seconds()
    elif (delta_in <= 0) & (delta_out >= 0):
        delta = (datetime.datetime.strptime(out_time_1, time_format) - datetime.datetime.strptime(in_time_1,
                                                                                                  time_format)).total_seconds()
    elif (delta_in >= 0) & (delta_out >= 0):
        delta = (datetime.datetime.strptime(out_time_1, time_format) - datetime.datetime.strptime(in_time_0,
                                                                                                  time_format)).total_seconds()
    elif (delta_in >= 0) & (delta_out <= 0):
        delta = (datetime.datetime.strptime(out_time_0, time_format) - datetime.datetime.strptime(in_time_0,
                                                                                                  time_format)).total_seconds()
    else:
        delta = 0

    return delta


def convert_to_sec(in_time, out_time, start, end):
    coef_in = []
    coef_out = []
    delta = []

    time_1_hour = en_check_time_1.get()
    time_1_hour_start = int(time_1_hour.split('-')[0].lstrip().rstrip())
    time_1_hour_end = int(time_1_hour.split('-')[1].lstrip().rstrip())

    time_2_hour = en_check_time_2.get()
    time_2_hour_start = int(time_2_hour.split('-')[0].lstrip().rstrip())
    time_2_hour_end = int(time_2_hour.split('-')[1].lstrip().rstrip())

    lesson_time = ((datetime.datetime.strptime(en_end.get(), time_format_in) - datetime.datetime.strptime(
        en_start.get(), time_format_in)).total_seconds()) / 60
    if (create_41.get()):
        if (lesson_time >= 110):
            start_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_2_hour_start)
            start_time = str(start_time)
            start_time = start_time.split(' ')[1]
            start_time = f"{start.split(' ')[0]} {start_time}" # время начала в формате 00:00
            end_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_2_hour_end)
            end_time = str(end_time)
            end_time = end_time.split(' ')[1]
            end_time = f"{end.split(' ')[0]} {end_time}"
        elif (lesson_time >=50):
            start_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_1_hour_start)
            start_time = str(start_time)
            start_time = start_time.split(' ')[1]
            start_time = f"{start.split(' ')[0]} {start_time}" # время начала в формате 00:00
            end_time = datetime.datetime.strptime(en_start.get(), time_format_in) + datetime.timedelta(minutes=time_1_hour_end)
            end_time = str(end_time)
            end_time = end_time.split(' ')[1]
            end_time = f"{end.split(' ')[0]} {end_time}"
    else:
        start_time = en_start.get()
        start_time = f"{start.split(' ')[0]} {start_time}:00"
        end_time = en_end.get()
        end_time = f"{end.split(' ')[0]} {end_time}:00"

    in_time = in_time
    out_time = out_time
    time_format = "%m/%d/%Y %H:%M:%S"  # %H:%M:%S     %m/%d/%Y %H:%M:%S

    # start = datetime.datetime.strptime(start, time_format)
    # end = datetime.datetime.strptime(end, time_format)
    start = datetime.datetime.strptime(start_time, time_format)
    end = datetime.datetime.strptime(end_time, time_format)
    print(start, end)

    for x, y in zip(in_time, out_time):  # поиск крайних значений посещения
        delta.append(
            (datetime.datetime.strptime(y, time_format) - datetime.datetime.strptime(x, time_format)).total_seconds())

    for x, y in zip(in_time, out_time):
        if float((start - datetime.datetime.strptime(x,
                                                     time_format)).total_seconds()) > 0:  # Если зашел раньше начала занятия
            if (datetime.datetime.strptime(y,
                                           time_format) - start).total_seconds() > 0:  # Если вышел после начала занятия
                coef_in.append(
                    (start - datetime.datetime.strptime(x, time_format)).total_seconds())  # Время началы - время входа
				# delta.append((datetime.datetime.strptime(y, "%m/%d/%Y %H:%M:%S") - datetime.datetime.strptime(x, "%m/%d/%Y %H:%M:%S")).total_seconds() - coef_in)
            else:  # если вышел до начала занятия
                coef_in.append((datetime.datetime.strptime(y, time_format) - datetime.datetime.strptime(x,
                                                                                                        time_format)).total_seconds())  # время входа минус время выхода
        else:
            coef_in.append(0)  # Если зашел после начала занятия

    for x, y in zip(in_time, out_time):
        if float((datetime.datetime.strptime(y,
                                             time_format) - end).total_seconds()) > 0:  # Если вышел после конца занятия
            if (end - datetime.datetime.strptime(x, time_format)).total_seconds() > 0:  # Если вошел до конца занятия
                coef_out.append(
                    (datetime.datetime.strptime(y, time_format) - end).total_seconds())  # Время выхода - время входа
            else:  # Если вошел после конца занятия
                coef_out.append((datetime.datetime.strptime(y, time_format) - datetime.datetime.strptime(x,
                                                                                                         time_format)).total_seconds())  # время входа минус время выхода
        else:
            coef_out.append(0)  # Вычисление входного коэффициента

    coef_const = [900 for x in range(0, len(coef_out))]

    df = pd.DataFrame({'Razn': delta, 'coef_out': coef_out, 'coef_in': coef_in, 'coef_const': coef_const})
    df['out'] = df['Razn'] - df['coef_out'] - df['coef_in']
    print(df['out'].tolist())
    return df['out'].tolist()


def similarityList(s1, *s2, shake=True):
    list_replace = []
    itog = False
    replaced = []
    list_0 = s2
    for numb, lis in enumerate(list_0):
        normalized1 = s1
        normalized2 = list_0[numb]
        for num2, y in enumerate(normalized2):
            matcher = fuzz.token_sort_ratio(normalized1, normalized2[num2])
            if (matcher >= 97) & (check_fio.get() == True):
                # print(matcher)
                # print(f'{normalized1}, {normalized2[num2]} - Совпадение')
                itog = True
                list_replace.append(s1)
                list_replace.append(normalized2[num2])
        if (itog == len(normalized1)) & (check_fio.get() == True):
            if (s1 != list_0[numb]):
                if (len(s1) > len(list_0[numb])) & (shake):
                    list_replace.append(s1)
                    list_replace.append(list_0[numb])
                elif (len(s1) <= len(list_0[numb])) & (shake):
                    list_replace.append(list_0[numb])
                    list_replace.append(s1)
                else:
                    list_replace.append(s1)
                    list_replace.append(list_0[numb])
            if (list_replace):
                print(list_replace)
                for number in range(1, len(list_replace)):
                    replaced.append(f'Замена {list_replace[number]} на {list_replace[0]}')
                    logger.insert(END, f'Замена {list_replace[number]} на {list_replace[0]}\n\n')
    if (list_replace):
        return list_replace
    else:
        return [" ", " "]


def find_intersections_same(dff1, start, end):
    uniq_name = dff1['ФИО']
    intersect_name = []
    intersect_time = []
    intersect_time_in_0 = []
    intersect_time_in_1 = []
    intersect_time_out_0 = []
    intersect_time_out_1 = []
    time_delta_list = []
    drop_row = []
    for name_num, name in enumerate(uniq_name):
        print(name)
        iters_for_name = len(dff1[dff1['ФИО'] == name]['ФИО'])
        for num in range(iters_for_name):
            in_time_0 = dff1[dff1['ФИО'] == name].iloc[num]['Время входа 0']
            out_time_0 = dff1[dff1['ФИО'] == name].iloc[num]['Время выхода 0']
            for num1 in range(iters_for_name):
                if (num < num1):
                    if (dff1[dff1['ФИО'] == name].iloc[num]['Время входа 0'] == dff1[dff1['ФИО'] == name].iloc[num1][
                        'Время входа 0']) & (dff1[dff1['ФИО'] == name].iloc[num]['Время выхода 0'] ==
                                             dff1[dff1['ФИО'] == name].iloc[num1]['Время выхода 0']):
                        in_time_1 = dff1[dff1['ФИО'] == name].iloc[num1]['Время входа 1']
                        out_time_1 = dff1[dff1['ФИО'] == name].iloc[num1]['Время выхода 1']
                    else:
                        in_time_1 = dff1[dff1['ФИО'] == name].iloc[num1]['Время входа 0']
                        out_time_1 = dff1[dff1['ФИО'] == name].iloc[num1]['Время выхода 0']
                    date_inersec = date_intersection(in_time_0, out_time_0, in_time_1, out_time_1)
                    if (date_inersec[0]):
                        drop_row.append(name_num)
                        intersect_time_in_0.append(in_time_0)
                        intersect_time_in_1.append(in_time_1)
                        intersect_time_out_0.append(out_time_0)
                        intersect_time_out_1.append(out_time_1)
                        intersect_name.append(name)
                        intersect_time.append(
                            f"Время входа {in_time_0} Время выхода {out_time_0}\nВремя входа {in_time_1} Время выхода {out_time_1}")
                        time_delta_list.append(time_delta(in_time_0, out_time_0, in_time_1, out_time_1, start, end))
    if (len(drop_row) > 0):
        drop_row = pd.Series(drop_row).unique()
        dff1 = dff1.drop(drop_row)
        intersec_tab = pd.DataFrame(
            {'ФИО': intersect_name, 'Время пересечения': intersect_time, 'Time Delta': time_delta_list,
             'Время входа 0': intersect_time_in_0, 'Время входа 1': intersect_time_in_1,
             'Время выхода 0': intersect_time_out_0, 'Время выхода 1': intersect_time_out_1})
        intersec_tab = intersec_tab.drop_duplicates()
        dff1 = intersec_tab.append(dff1)
        dff1 = dff1.drop_duplicates()
        try:
            dff1 = dff1.drop('level_0', axis=1)
        except:
            pass
        dff1 = dff1.reset_index()
        return dff1
    else:
        return dff1


def convert():
    try:
        summ = df1.append(df2).append(df3).append(df4).append(df5).append(df6).append(df7).append(df8).append(
            df9).append(df10)
    except:
        logger.insert(END, f'ERROR: Ошибка!!!\n')

    dict_time = dict(zip(name_list_all, time_list_all))
    print(dict_time)

    fio = []
    for x in summ['Имя (настоящее имя)'].tolist():
        fio.append(x.split(' (')[0])

    summ['ФИО'] = fio
    summ['Продолжительность (минуты)'] = summ['Продолжительность (минуты)'].astype('float64')

    for x in list(summ['ФИО'].tolist()):  #### HERE BUGGGGGG!!!!!
        ls = similarityList(x, summ['ФИО'].tolist(), shake=True)
        for y in range(1, len(ls)):
            summ['ФИО'] = summ['ФИО'].replace(ls[y], ls[0])

    for x in list(df0['ФИО гражданина'].tolist()):
        ls = similarityList(x, summ['ФИО'].tolist(), shake=False)
        print(f'Получаем {ls}')
        for y in range(1, len(ls)):
            summ['ФИО'] = summ['ФИО'].replace(ls[y], ls[0])

	# Поиск совпадений из оставшихся позиций улучшенным алгоритмом
    ost_df = []
    for num0, x in enumerate(summ['ФИО'].tolist()):
        if x not in (df0['ФИО гражданина'].tolist()):
            ost_df.append(x)  # люди которые не попали в список

    for name_ost in ost_df:
        matcher = 0
        matcher_ = 0
        name_ost_list = name_ost.split(' ')
        name_ost_ = name_ost
        name_ost_ = name_ost_.replace(' ', '')
        for good_name in df0['ФИО гражданина'].tolist():
            good_name_list = good_name.split(' ')
            good_name_ = good_name
            good_name_.replace(' ', '')
            matcher = fuzz.token_sort_ratio(list(name_ost_), list(good_name_))
            if (matcher > 97) & (check_fio.get() == True):  # Поиск обьедененных слов в ФИО
                summ['ФИО'] = summ['ФИО'].replace(name_ost, good_name)
                logger.insert(END, f'\n')
                logger.insert(END, f'Замена {name_ost} на {good_name} (Причина: Возможно слитное написание ФИО):\n')
            elif (matcher > 92) & (check_fio.get() == True):
                matcher = fuzz.token_sort_ratio((name_ost), (good_name))
                print(name_ost, good_name, matcher)
                if (matcher > 94):
                    summ['ФИО'] = summ['ФИО'].replace(name_ost, good_name)
                    logger.insert(END, f'\n')
                    logger.insert(END, f'Замена {name_ost} на {good_name}: (Причина: Возможно грубая опечатка в ФИО)\n')
    ##############################################################

    summ = summ.reset_index()
	# Добавляем новую колнку с датой занятия в формате день.месяц.год[2]
    summ['Date_zaniat'] = convert_summ_date(summ['Name'].tolist())
    # summ.to_excel('summ.xlsx')
	# Выводим информацию о всех уникальных участниках конференции
    tab_analize = summ[['ФИО', 'Name']]
    tab_analize = tab_analize.drop_duplicates(subset=['ФИО'])
    tab_analize = tab_analize.reset_index()
    tab_analize_to_drop = []
    for num_tab_analize_fio, tab_analize_fio in enumerate(tab_analize['ФИО'].tolist()):
        for not_used_name in list_of_not_used:
            if not_used_name in tab_analize_fio.lower():
                print(f'Из уникальных участников исключен {tab_analize_fio}')
                tab_analize_to_drop.append(num_tab_analize_fio)
    tab_analize_to_drop = list(set(tab_analize_to_drop))
    tab_analize = tab_analize.drop(tab_analize_to_drop)
    if (uniq_create.get()):
        tab_analize.to_excel(f"{group_name}_uniq_parti.xlsx")
        logger.insert(END, f'\n\nСоздан файл {group_name}_uniq_parti.xlsx уникальных участников\n')
    uniq_analize = summ['Name'].unique()
    logger.insert(END, "\nКоличество уникальных участников на занятии:\n")
    for x in uniq_analize:
        logger.insert(END, f"{x} -------> {len(tab_analize[tab_analize['Name'] == x])}\n")
    uniq_analize = []

    new_df = []
    new_df_prefix = []
    for num0, x in enumerate(summ['ФИО'].tolist()):
        if x not in (df0['ФИО гражданина'].tolist()):
            print(f'Не в списке - {x}')
            new_df.append(num0)
            if (len(re.findall(r'куратор', x.lower())) > 0):
                new_df_prefix.append(num0)
        else:
            new_df_prefix.append(num0)

    summ_prefix = summ.drop(new_df_prefix)
    summ_prefix.reset_index()
	# summ_prefix.drop_duplicates(subset = ['ФИО'])
    summ_prefix = summ_prefix[['ФИО', 'Secs', 'Name', 'Time']]
	# summ_prefix.to_excel('summ_prefix.xlsx')

    logger.insert(END, f'\n')
    logger.insert(END, f'Список людей не включенных в основную ведомость:\n')
    for number_not in range(len(summ_prefix['ФИО'].tolist())):
        logger.insert(END, "\n")
        logger.insert(END, summ_prefix.iloc[number_not]['ФИО'])
        logger.insert(END, "\n")
        logger.insert(END, summ_prefix.iloc[number_not]['Name'])
        logger.insert(END, "\n ---------------------------------------")
    logger.insert(END, f'\n')
    logger.insert(END,
                  f"Количество уникальных участников не включенных в ведомость: {len(pd.Series(summ_prefix['ФИО'].tolist()).unique())}")
    summ = summ.drop(new_df)

    # Сравниваем дату проведения занятия и время зачисления участника. Если участник был зачислен позже чем был
    # на занятии, то обнуляем его поле 'Secs' и Минуты проведенные на занятии
    new_secs = []
    for num_name, fio_summ in enumerate(summ['ФИО'].tolist()):
        if fio_summ in df0['ФИО гражданина'].tolist():
            for num_name1, fio_summ1 in enumerate(df0['ФИО гражданина'].tolist()):
                if fio_summ == fio_summ1:
                    if (time_equal(df0.iloc[num_name1]['Дата зачисления'], summ.iloc[num_name]['Date_zaniat'])):
                        new_secs.append(summ.iloc[num_name]['Secs'])
                    else:
                        logger.insert(END,
                                      f"\n{summ.iloc[num_name]['ФИО']} - был(а) на занятии раньше зачисления\nБыла: {summ.iloc[num_name]['Date_zaniat']} Дата зачисления: {df0.iloc[num_name1]['Дата зачисления']}\n")
                        new_secs.append(0)
        else:
            new_secs.append(summ.iloc[num_name]['Secs'])
    print(len(new_secs))
    print(len(summ['Secs'].tolist()))
    summ['Secs'] = new_secs
    ############################################

    print(new_df)

    not_in_list = []
    for num0, x in enumerate(df0['ФИО гражданина'].tolist()):
        if x not in (summ['ФИО'].tolist()):
            not_in_list.append(x)

    summ = summ.append(pd.DataFrame({'ФИО': not_in_list, 'Name': [df1_name for x in range(len(not_in_list))],
                                     'Продолжительность (минуты)': [0 for x in range(len(not_in_list))]}))
    if (create_41.get()):
        add_to_secs = []
        for duration_minutes in summ['Продолжительность (минуты)']:
            try:
                if duration_minutes > 0:
                    add_to_secs.append(0.0001)
                else:
                    add_to_secs.append(0)
            except:
                add_to_secs.append(0)
        summ['add_to_secs'] = add_to_secs
        summ['Secs'] = summ['Secs'] + summ['add_to_secs']
    #summ.to_excel('summ.xlsx')

    tell_phone = []
    print(len(summ['ФИО'].tolist()))
    for numb, fio in enumerate(summ['ФИО'].tolist()):
        if fio in df0['ФИО гражданина'].tolist():
            for numb0, fio0 in enumerate(df0['ФИО гражданина'].tolist()):
                if (fio == fio0):
                    print(df0.iloc[numb0]['Контактный телефон гражданина'])
                    tell_phone.append(df0.iloc[numb0]['Контактный телефон гражданина'])
                else:
                    continue
        else:
            tell_phone.append('-')
        print(len(tell_phone))
        print(numb)
        print(summ['ФИО'].tolist()[numb])

    try:
        summ['Номер телефона'] = tell_phone
    except:
        print('В исходной ведомости есть одинаковые ФИО! Проверьте ведомость!')
        # summ.to_excel('summ.xlsx')
    summ_ = summ.copy()
    summ_['Start_time'] = summ_['Name'].apply(lambda x: x.split('Время проведения занятия')[1].split('-')[0].replace('(', '').replace('\n', ''))
    summ_['End_time'] = summ_['Name'].apply(lambda x: x.split('Время проведения занятия')[1].split('-')[1].replace(')', ''))
    In_time_list = []
    for num, x in enumerate(summ_['Время входа'].tolist()):
        try:
            In_time_list.append(summ_.iloc[num]['Время входа'].split(' ')[1])
        except:
            In_time_list.append(0)
    summ_['In_time'] = In_time_list
    Out_time_list = []
    for num, x in enumerate(summ_['Время выхода'].tolist()):
        try:
            Out_time_list.append(summ_.iloc[num]['Время выхода'].split(' ')[1])
        except:
            Out_time_list.append(0)
    summ_['Out_time'] = Out_time_list
    #Ищем опоздавших более чем на 15 минут
    summ_appended = summ_[summ_['In_time'] == 0]
    summ_ = summ_[summ_['In_time'] != 0]
    After_15_list = []
    for num, x in enumerate(summ_['ФИО'].tolist()):
        try:
            secs_after = (datetime.datetime.strptime(summ_.iloc[num]['In_time'], time_format_in_secs) - datetime.datetime.strptime(summ_.iloc[num]['Start_time'], time_format_in)).total_seconds()
            if(secs_after > 900):
                After_15_list.append(True)
            else:
                After_15_list.append(False)
        except:
            After_15_list.append(False)
    summ_['After_15'] = After_15_list
    After_15_list = []
    columns_summ = summ_['Name'].unique()
    #summ_.to_excel('summ_.xlsx')
    for y in columns_summ:
        summ_i = summ_[(summ_['Name'] == y)]
        for num, x in enumerate(summ_i['ФИО'].tolist()):
            After_15_list.append(summ_i[(summ_i['ФИО'] == x)]['After_15'].all())
    summ_['After_15'] = After_15_list
    cashed_names = []
    enter_status = []
    for num, x in enumerate(summ_['ФИО'].tolist()):
        if (summ_.iloc[num]['After_15']):
            if (x not in cashed_names):
                cashed_names.append(x)
                enter_status.append(0.01)
            else:
                enter_status.append(0)
        else:
            enter_status.append(0)
    summ_['Enter_status_after'] = enter_status

    #Ищем тех кто вышел за 15 минут до окончания занятия
    Before_15_list = []
    for num, x in enumerate(summ_['ФИО'].tolist()):
        try:
            secs_before = (datetime.datetime.strptime(summ_.iloc[num]['End_time'], time_format_in) - datetime.datetime.strptime(summ_.iloc[num]['Out_time'],
                                                     time_format_in_secs)).total_seconds()
            if (secs_before > 900):
                Before_15_list.append(True)
            else:
                Before_15_list.append(False)
        except:
            Before_15_list.append(False)
    summ_['Before_15'] = Before_15_list
    Before_15_list = []

    for y in columns_summ:
        summ_i = summ_[(summ_['Name'] == y)]
        for num, x in enumerate(summ_i['ФИО'].tolist()):
            Before_15_list.append(summ_i[(summ_i['ФИО'] == x)]['Before_15'].all())
    summ_['Before_15'] = Before_15_list
    cashed_names = []
    enter_status = []
    for num, x in enumerate(summ_['ФИО'].tolist()):
        if (summ_.iloc[num]['Before_15']):
            if (x not in cashed_names):
                cashed_names.append(x)
                enter_status.append(0.03)
            else:
                enter_status.append(0)
        else:
            enter_status.append(0)
    summ_['Enter_status_before'] = enter_status

    summ_['Secs'] = summ_['Secs'] + summ_['Enter_status_after'] + summ_['Enter_status_before']
    summ_ = summ_.append(summ_appended)
    summ = summ_.copy()
    table_prefix = pd.pivot_table(summ_prefix, index=['ФИО'], columns='Name', values='Secs',
                                  aggfunc=np.sum)  # aggfunc = np.num
    table = pd.pivot_table(summ, index=['ФИО', 'Номер телефона'], columns=['Name'], values='Secs',
                           aggfunc=np.sum)  # aggfunc = np.num

    table.to_excel(f'{group_name}_seconds.xlsx')
    #table_prefix.to_excel('table_prefix.xlsx')
	# table.replace(np.nan, '-')
    table['Всего посещений'] = table.sum(axis=1)
    table['Всего посещений'] = round(table['Всего посещений'] / 60 / 60, 0)
    # Подсчет часов занятия

    min_secs_105 = 60 * 105
    min_secs_45 = 60 * 45
    min_secs_60 = 60 * 60

    time_1_hour = en_check_time_1.get()
    time_1_hour_start = int(time_1_hour.split('-')[0].lstrip().rstrip())
    time_1_hour_end = int(time_1_hour.split('-')[1].lstrip().rstrip())

    time_2_hour = en_check_time_2.get()
    time_2_hour_start = int(time_2_hour.split('-')[0].lstrip().rstrip())
    time_2_hour_end = int(time_2_hour.split('-')[1].lstrip().rstrip())
    min_secs_41 = (time_1_hour_end - time_1_hour_start) * 60
    min_secs_101 = (time_2_hour_end - time_2_hour_start) * 60

    print(min_secs_41, min_secs_101)
    group_time = 0  # тип группы часовая или двух часовая

    sum_plus_list = []
    summ_col = []
    if(create_41.get()):
        for x in table.columns[:-1]:
            zn0 = []
            for zn in table[x].tolist():
                if (isinstance(zn, float)):
                    if (zn == 0):
                        zn0.append('н')
                    elif (dict_time[x] >= 110):
                        group_time = 110
                        if (zn >= 1): #if (zn >= min_secs_101)
                            zn0.append('+')
                        #elif (zn > 0):
                            #zn0.append('б')
                        else:
                            zn0.append('н')
                    elif (dict_time[x] >= 50):
                        group_time = 50
                        if (zn >= 1): #(zn >= min_secs_41):
                            zn0.append('+')
                        #elif (zn > 0):
                            #zn0.append('б')
                        else:
                            zn0.append('н')
                else:
                    zn0.append(zn)
            table[x] = zn0
            sum_plus_list.append(x)  # Названия колонок
            r = table[x].tolist()
            summ_col.append(np.sum(np.array(r) == '+'))
    else:
        for x in table.columns[:-1]:
            zn0 = []
            for zn in table[x].tolist():
                if (isinstance(zn, float)):
                    if (zn == 0):
                        zn0.append('н')
                    elif (dict_time[x] >= 110):
                        group_time = 110
                        if (zn >= min_secs_105):
                            zn0.append('+')
                        elif (zn >= min_secs_60):
                            enter_st = round(math.modf(zn)[0], 2)
                            #print(enter_st)
                            if(enter_st == 0.01):
                                zn0.append('о')
                            elif(enter_st == 0.03):
                                zn0.append('у')
                            elif(enter_st == 0.04):
                                zn0.append('оу')
                            else:
                                zn0.append('б')
                        elif (zn > 0.1):
                            zn0.append('б')
                        else:
                            zn0.append('н')
                    elif (dict_time[x] >= 50):
                        group_time = 50
                        if (zn >= min_secs_45):
                            zn0.append('+')
                        elif (zn > 0.1):
                            enter_st = round(math.modf(zn)[0], 2)
                            if (enter_st == 0.01):
                                zn0.append('о')
                            elif(enter_st == 0.03):
                                zn0.append('у')
                            else:
                                zn0.append('б')
                        else:
                            zn0.append('н')
                else:
                    zn0.append(zn)
            table[x] = zn0
            sum_plus_list.append(x)  # Названия колонок
            r = table[x].tolist()
            summ_col.append(np.sum(np.array(r) == '+'))
    # summ_col = summ_col.append('-')
    data_row = pd.DataFrame(dict(zip(sum_plus_list, summ_col)), index=[0])

    print(summ_col)
    print(sum_plus_list)
    table = table.reset_index()

    table = table.append(data_row, ignore_index=True)
	# table.iloc[-2] = table.iloc[-2].replace(0, '-')

    for x in table_prefix.columns:
        zn1 = []
        for num, zn in enumerate(table_prefix[x].tolist()):
            if (isinstance(zn, float)):
                if (zn == 0):
                    zn1.append(0)
                elif (dict_time[x] >= 110):
                    if (zn >= 1):#(zn >= min_secs_101)
                        zn1.append(1)
                    else:
                        zn1.append(0)
                elif (dict_time[x] >= 50):
                    if (zn >= 1):#(zn >= min_secs_41)
                        zn1.append(1)
                    else:
                        zn1.append(0)
            else:
                zn1.append(zn)
        table_prefix[x] = zn1
    table_prefix['sum'] = table_prefix.sum(axis=1)
	# table_prefix.to_excel('table_prefix.xlsx')
    table_prefix = table_prefix[table_prefix['sum'] > 0]
    table_prefix = table_prefix.replace(1, '+')
    table_prefix = table_prefix.replace(0, 'н')
    table_prefix = table_prefix.drop(columns=['sum'])

    # Суммирование + по строкам
    summ_poseh = []
    for x in range(len(table['ФИО'].tolist())):
        r = table.iloc[x].values
        summ_poseh.append(np.sum(r == '+'))
    table['Всего часов занятий'] = summ_poseh

    if (group_time == 110):
        table['Всего посещений'] = table['Всего часов занятий'] * 2  # Переформатируем ьаблицу по желанию заказчика
    elif (group_time == 50):
        table['Всего посещений'] = table['Всего часов занятий'] * 1  # Переформатируем ьаблицу по желанию заказчика

    # table['Всего часов занятий'] = round(table['Всего часов занятий']/60/60, 0)
    table[['Всего посещений', 'Всего часов занятий']] = table[['Всего часов занятий', 'Всего посещений']]

    pesehenie = table['Всего посещений'].sum()
    chasov = table['Всего часов занятий'].sum()  # Не используется
	# chasov = table['Всего часов занятий'].sum()

    table = table.replace(np.nan, '-')
    df0['ФИО гражданина'] = df0['ФИО гражданина'].str.replace('\xa0', ' ')
    table.index = table['ФИО']
    df0.index = df0['ФИО гражданина']
    table = table.align(df0, join='right', axis=0)[0]
    table.index = table[table.columns[0]]
    del (table[table.columns[0]])
    table = table.reset_index()

	# table = table.reset_index()
	# table.index = [x for x in range(1, len(table.values)+1)]

    rows_ = len(table.values)
    cols_ = len(table.columns)

    table = table.append(pd.Series(), ignore_index=True).append(pd.Series(), ignore_index=True).append(
        table_prefix.reset_index())

	# table = table.reset_index()
    table.index = [x for x in range(1, len(table.values) + 1)]

	# rename columns для того чтобы они вмещались в колонку
    name_sum_zan = f'''Всего
часов
занятий'''

    name_sum_poseh = f'''Всего
посещений'''

    table.rename(columns={'Всего посещений': name_sum_poseh, 'Всего часов занятий': name_sum_zan}, inplace=True)

    table.to_excel(f'{group_name}.xlsx', startrow=12)
    rows = len(table.values)
    cols = len(table.columns)

    A2_str = f'{group_name}; {month_day}, {year} г.'
    A4_str = 'Ведомость учета посещаемости гражданами площадок активности'
    A5_str = 'в рамках реализации проекта «Московское долголетие»'
    A8_str = f'Месяц, год проведения занятия: {month_day} {year} г.'
    A9_str = 'Организация: ИП Артемова Юлия Сергеевна'
    A10_str = 'Соглашение от «31» декабря 2019 г. № 1288800702020'
    A11_str = f'Наименование занятия "Здорово жить"   Код группы: {group_name}'

    my_wb = openpyxl.load_workbook(f'{group_name}.xlsx')
	# my_sheet = my_wb.get_sheet_by_name('Sheet1')
    my_sheet = my_wb['Sheet1']
    A2 = my_sheet['A2']
    A2.value = A2_str
    A4 = my_sheet['A4']
    A4.value = A4_str
    A5 = my_sheet['A5']
    A5.value = A5_str
    A8 = my_sheet['A8']
    A8.value = A8_str
    A9 = my_sheet['A9']
    A9.value = A9_str
    A10 = my_sheet['A10']
    A10.value = A10_str
    A11 = my_sheet['A11']
    A11.value = A11_str

    pesehenie_set = my_sheet.cell(column=cols_, row=rows_ + 14)
    pesehenie_set.value = pesehenie
    chasov_set = my_sheet.cell(column=cols_ + 1, row=rows_ + 14)
	# chasov_set.value = chasov
    if (group_time == 110):
        chasov_set.value = pesehenie * 2
    elif (group_time == 50):
        chasov_set.value = pesehenie * 1

    tail_str_0 = 'ИП Артемова Юлия Сергеевна'
    tail_str_1 = 'Представитель ИП Артемова Ю.С.'
    tail_str_2 = 'По доверенности 1-06/20 от 30.06.2020'
    tail_str_3 = 'Шкадина С.С'
    tail_str_4 = f'«___» ________________ {year} г.'

    tail_str_5 = f'''* При присутствии участника группы на занятии ставится знак «+», при отсутствии - буква «Н»,'''
    tail_str_6 = f'''при этом если участник группы не участвовал в занятии более 15 минут, ставится буква «Н»'''
    tail_str_7 = f'''**Ведомость должна быть заполнена до окончания занятия, в случае отсутствия гражданина'''
    tail_str_8 = f'''на занятии в ячейке проставляется буква «Н»'''

    tail_str_0_s = my_sheet.cell(column=2, row=10 + rows + 7)
    tail_str_0_s.value = tail_str_0
    tail_str_1_s = my_sheet.cell(column=2, row=10 + rows + 8)
    tail_str_1_s.value = tail_str_1
    tail_str_2_s = my_sheet.cell(column=2, row=10 + rows + 9)
    tail_str_2_s.value = tail_str_2
    tail_str_3_s = my_sheet.cell(column=2, row=10 + rows + 10)
    tail_str_3_s.value = tail_str_3
    tail_str_4_s = my_sheet.cell(column=2, row=10 + rows + 11)
    tail_str_4_s.value = tail_str_4
    tail_str_5_s = my_sheet.cell(column=2, row=10 + rows + 14)
    tail_str_5_s.value = tail_str_5
    tail_str_6_s = my_sheet.cell(column=2, row=10 + rows + 15)
    tail_str_6_s.value = tail_str_6
    tail_str_7_s = my_sheet.cell(column=2, row=10 + rows + 16)
    tail_str_7_s.value = tail_str_7
    tail_str_8_s = my_sheet.cell(column=2, row=10 + rows + 17)
    tail_str_8_s.value = tail_str_8

	# Установка ширины постоянных колонок
    my_sheet.column_dimensions['A'].width = 5
    my_sheet.column_dimensions['B'].width = 37
    my_sheet.column_dimensions['C'].width = 17
    my_sheet.row_dimensions[13].height = 96

	# Установка ширины переменных колонок
    cols_dict = {0: 'D', 1: 'E', 2: 'F', 3: 'G', 4: 'H', 5: 'I', 6: 'J', 7: 'K', 8: 'L', 9: 'M'}
    for x in range(cols):
        my_sheet.column_dimensions[cols_dict[x]].width = 12
        for y in range(rows + 15):
            # print(f'{cols_dict[x]}{y+12}')
            currentRow = my_sheet[f'{cols_dict[x]}{y + 14}']  # or currentCell = ws['A1']
            currentRow.alignment = openpyxl.styles.Alignment(horizontal='center')

	# Установка Шапки
    my_sheet.merge_cells('A4:F4')
    my_sheet.merge_cells('A5:F5')
    currentCell = my_sheet['A4']
    currentCell.font = openpyxl.styles.Font(size=16)
    currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')
    currentCell = my_sheet['A5']  # or currentCell = ws['A1']
    currentCell.font = openpyxl.styles.Font(size=16)
    currentCell.alignment = openpyxl.styles.Alignment(horizontal='center')

    my_wb.save(f"{group_name}.xlsx")

	# Запись данных в БД sql reg.ru
    try:
        pcname = os.getenv('COMPUTERNAME')
        mydb = mysql.connector.connect(
            host="server224.hosting.reg.ru",
            user="u1051830_alexand",
            password="Megare926",
            port="3306",
            db='u1051830_dolgoletie_stats')
        mycursor = mydb.cursor()
        mycursor.execute(f"INSERT INTO dolgoletie VALUES (NULL, '{datetime.datetime.now()}', '{pcname}_{create_41.get()}_{group_name}')")
        mydb.commit()
    # mycursor.execute("SELECT * FROM dolgoletie_test")

    # for x in mycursor:
    #	print(x)
    except:
        pass

    # logger.insert(END, f'ERROR: Файл не создан, произошла ошибка!\n')


open_0_file = Button(f_left, text='Select a .xlsx(Excel) file', command=file_opener_0)
open_1_file = Button(f_left, text='Select a .csv file', command=file_opener_1)
open_2_file = Button(f_left, text='Select a .csv file', command=file_opener_2)
open_3_file = Button(f_left, text='Select a .csv file', command=file_opener_3)
open_4_file = Button(f_left, text='Select a .csv file', command=file_opener_4)
open_5_file = Button(f_left, text='Select a .csv file', command=file_opener_5)
open_6_file = Button(f_left, text='Select a .csv file', command=file_opener_6)
open_7_file = Button(f_left, text='Select a .csv file', command=file_opener_7)
open_8_file = Button(f_left, text='Select a .csv file', command=file_opener_8)
open_9_file = Button(f_left, text='Select a .csv file', command=file_opener_9)
open_10_file = Button(f_left, text='Select a .csv file', command=file_opener_10)
l0 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l1 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l2 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l3 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l4 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l5 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l6 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l7 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l8 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l9 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l10 = Label(f_left, fg="green", height=1, font=("Courier", 10))
l_select = Label(f_left, text='Выберите файл ведомости:')
l_select_csv = Label(f_left, text='Выберите файлы отчета ZOOM:')
l_start_time = Label(f_left, text='Введите время начала занятия группы:')
l_end_time = Label(f_left, text='Введите время окончания занятия группы:')
l_group_name = Label(f_left, text='Группа')
create = Button(f_left, text='Создать ведомость', command=convert)
reset = Button(f_left, text='Сброс', command=reset_functions)
en_start = Entry(f_left, width=5)
en_start.insert(0, '12:00')
en_end = Entry(f_left, width=5)
en_end.insert(0, '14:00')
en_check_time_1 = Entry(f_left, width = 7)
en_check_time_2 = Entry(f_left, width = 7)
en_check_time_1.insert(0, '41-60')
en_check_time_2.insert(0, '101-120')
####################################################################################
f_left.pack(side=LEFT)
f_right.pack(side=RIGHT)
l_group_name.pack()
l_start_time.pack()
en_start.pack()
l_end_time.pack()
en_end.pack()
Label(f_left).pack()
checkbutton.pack()
checkbutton_create_uniq.pack()
checkbutton_41_60.pack()
en_check_time_1.pack()
en_check_time_2.pack()
# Label().pack()
l_select.pack()
Label().pack()
open_0_file.pack()
l0.pack()
create.pack()  ###############################
Label(f_left).pack()  ########################
reset.pack()  ################################
l_select_csv.pack()
open_1_file.pack()
l1.pack()
open_2_file.pack()
l2.pack()
open_3_file.pack()
l3.pack()
open_4_file.pack()
l4.pack()
open_5_file.pack()
l5.pack()
open_6_file.pack()
l6.pack()
open_7_file.pack()
l7.pack()
open_8_file.pack()
l8.pack()
open_9_file.pack()
l9.pack()
open_10_file.pack()
l10.pack()
# create.pack()
# Label(f_left).pack()
# reset.pack()
logger.pack()
logger.insert(END, 'Перед добавлением файлов введите время занятия группы!\n')
mainloop()
